import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import platform
import re
from datetime import datetime
from colorama import Fore, init
from typing import Tuple, Dict
init(autoreset=True)
composer_deps = []
npm_deps = []
json_data = {}
def extract_file_deps(file: str, _case: str) -> None:
    """
		Extracts dependencies from a configuration file according to its filename
		:param str file: the file path
		:param str _case: the filename pattern
		:return None
    """
    def evaluate_version(d: str, exist_v: str, load_v: str) -> bool:
      l_n = re.sub(r'[^0-9\.]', '', load_v) if len(d) > 3 else ''
      if (not e_n and l_n) or (exist_v == '*' and l_n):
        return True
      else:
        e_n = re.sub(r'[^0-9\.]', '', exist_v) if len(d) > 3 else ''
        if not l_n:
          return False
        if not e_n:
          return True
        else:
          e_v_parts = [int(p) if p.isdigit() else 0 for p in exist_v.split('.')]
          l_v_parts = [int(p) if p.isdigit() else 0 for p in load_v.split('.')]
          len_e_parts = len(e_v_parts)
          len_l_parts = len(l_v_parts)
          if ((len_e_parts == 1 or len_e_parts == 2) and (len_l_parts == 1 or len_l_parts == 2)) and float(e_n) < float(l_n):
            for i, _ in enumerate(l_v_parts):
              e_part = e_v_parts[i] if i < len_e_parts else 0
              l_part = l_v_parts[i] if i < len_l_parts else 0
              return True if isinstance(e_part, int) and isinstance(l_part, int) else False
          elif len_l_parts > 2:
            return True
          else:
            return False
    def case_append(item: Tuple[str]) -> None:
        print(Fore.WHITE + f'Got {item[2]} version {item[3]} as a dependency')
        if _case == 'npm':
            npm_deps.append(item)
        elif _case == 'composer':
            composer_deps.append(item)
    def case_remove(found_in_list: object) -> None:
        if _case == 'npm':
            npm_deps.remove(found_in_list)
        elif _case == 'composer':
            composer_deps.remove(found_in_list)
    try:
        rel = os.path.relpath(file)
        with open(file, 'r', encoding='utf-8') as f:
            content = json.load(f)
        def add_deps(c: dict, prop: str, group: str = 'general') -> None:
            """
						Adds dependencies in a list according to the key string
						:param dict c: the content (key:value pairs) to be included
						:param 'dependencies'|'devDependencies'|'require'|'require-dev' prop: the literal key to be read
						:param 'general'|'dev' group: the type of dependency key
						:return None
            """
            for d, v in c.get(prop, {}).items():
                found_in_list = None
                if _case == 'composer':
                  append_json(d, loaded_v=v, _case='composer')
                  for i_c in composer_deps[:]:
                        if i_c[2] == d:
                            found_in_list = i_c
                            break
                elif _case == 'npm':
                  append_json(d, loaded_v=v, _case='npm')
                  for i_n in npm_deps[:]:
                        if i_n[2] == d:
                            found_in_list = i_n
                            break
                else:
                    raise ValueError('Case does not exist in dependency check switch.')
                if found_in_list:
                    existing_v = found_in_list[3]
                    if evaluate_version(d, exist_v=existing_v, loaded_v=v):
                      case_remove(found_in_list)
                      case_append((_case, group, d, v, rel))
                    else:
                      pass
                else:
                    case_append((_case, group, d, v, rel))
        def append_json(d: str, loaded_v: str, _case: str, group: str = 'general') -> None:
          """
          Updates the json listing the dependencies found in each of the files
          :param str _case: the filename pattern
          :param str d: the dependency name
          :param str v: the dependency version
					:param 'general'|'dev' group: the type of dependency key
					:return None
          """
          if _case not in json_data:
            json_data[_case] = {}
          if group not in json_data[_case]:
            json_data[_case][group] = {}
          if rel not in json_data[_case][group]:
            json_data[_case][group][rel] = {}
          g = json_data[_case][group][rel]
          for dep in g:
            if d in dep:
              if evaluate_version(d, exist_v=g[d], loaded_v=loaded_v): g[d] = loaded_v
              else: return
            else: g[d] = loaded_v
        if _case == 'npm':
            if 'dependencies' in content:
                add_deps(content, prop='dependencies')
            if 'devDependencies' in content:
                add_deps(content, prop='devDependencies', group='dev')
        elif _case == 'composer':
            if 'require' in content:
                add_deps(content, prop='require')
            if 'require-dev' in content:
                add_deps(content, prop='require-dev', group='dev')
        else:
            raise ValueError('Invalid given case.')
    except FileNotFoundError as e:
        print(f'The passed file {file} could not be recognized as a file: {e}')
    except ValueError as e:
        if not isinstance(file, str) or not isinstance(_case, str):
            print(Fore.RED + f"""Unexpected argument type given:
                                Type of file path: {type(file)}
                                Type of case: {type(_case)}""")
        elif _case not in ('npm', 'composer'):
            print(Fore.RED + f'The passed case for evaluation is not valid.\nObtained value: {_case}')
        else:
            print(Fore.RED + f'There has been an undefined error while working with a function value: {e}')
    except PermissionError:
        print(Fore.RED + f'Permission for processing {file} denied by {os.name} on {platform.platform()}')
    except Exception as e:
        print(Fore.RED + f'As unknown error has occurred: {e}')
def extract_all_deps() -> bool:
    """
    Reads recursively the dependencies listed in the config files for package managers in the directory
    :return bool defining if the operation was successful or not
    """
    print(Fore.YELLOW + f'Running extraction of dependencies...')
    try:
        for r, _, fs in os.walk(os.path.dirname(os.path.abspath(__file__))):
            for f in fs:
                full_path = os.path.join(r, f)
                if f == 'composer.json' or f == 'package.json':
                    print(Fore.BLUE + f'Extracting dependencies from {full_path}...')
                    if f == 'composer.json':
                        extract_file_deps(full_path, 'composer')
                    elif f == 'package.json':
                        extract_file_deps(full_path, 'npm')
        name = f'extracted_dependencies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        all_deps = composer_deps + npm_deps
        if all_deps:
            pd.DataFrame(all_deps, columns=('Case', 'Group', 'Dependency', 'Version', 'Related Path')).to_excel(name, index=False, sheet_name="MainList")
            xls_path = os.path.abspath(name)
            if (os.path.exists(xls_path) and os.path.isfile(xls_path)):
                wb = load_workbook(xls_path)
                ws = wb.active
                header_fill = PatternFill(start_color="FFFDD0", end_color="FFFDD0", fill_type="solid")
                body_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    for cell in row:
                        if cell.value == '*':
                            cell.font = Font(color='F79646', bold=True)
                        elif isinstance(cell.value, str) and re.search(r'\|', cell.value):
                            cell.font = Font(color='1F497D', italic=True)
                        if idx % 2 == 0:
                            cell.fill = body_fill
                wb.save(xls_path)
                print(Fore.GREEN + f'Successfully extracted and tabled current directory dependencies.\nYou can find it in {name}')   
            else:
                raise FileNotFoundError(f'Could not write the file {name}')
        else:
            print(Fore.YELLOW + 'No dependencies found in the working directory.')
        return True
    except PermissionError:
        print(Fore.RED + f'Processing of {os.getcwd()} denied by {os.name} on {platform.platform()}')
        return False
    except Exception as e:
        print(Fore.RED + f'An unknown error has stopped the execution: {e}')
        return False

if __name__ == '__main__':
    extract_all_deps()