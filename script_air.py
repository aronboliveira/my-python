import openpyxl
import re
import os
import sys
from colorama import Fore, init
init(autoreset=True)
def iterate(f):
  workbook = openpyxl.load_workbook(f)
  for sh_n in workbook.sheetnames:
    sh = workbook[sh_n]
    cols = range(1, sh.max_column + 1)
    print(Fore.YELLOW + f'Processing {sh_n}...')
    hs = [sh.cell(row=1, column=col).value for col in cols]
    targ_cols = {}
    for i, h in enumerate(hs, start=1):
      if h:
        str_h = str(h).lower().strip()
      if '/' in str_h: str_h = str_h.split('/')[0]
      if str_h not in ['rota', 'origem', 'destino', 'km']: continue
      else: 
        print(Fore.GREEN + f'Validated header: {str_h}')
        targ_cols[str_h] = i
    print(Fore.YELLOW + f'Selected headers: indexes {str(targ_cols)}')
    if 'rota' not in targ_cols: 
      print(Fore.RED + 'Error accessing rota. Aborting process.')
      return
    rota_i = targ_cols['rota']
    if 'origem' in targ_cols and 'destino' in targ_cols:
      print(Fore.YELLOW + 'Reading route segments...')
      for j, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
        rota_v = row[rota_i - 1]
        if rota_v is None or not isinstance(rota_v, str):
          print(Fore.RED + 'Error validating value for rota')
          continue
        parts = re.split(r'/(?=[A-Za-z])', rota_v)
        if len(parts) < 2: continue
        origem = parts[0].strip()[:3]
        if '*' in origem: destino = destino.split('*')[0].strip()[:3]
        destino = parts[1].strip()[:3]
        sh.cell(row=j, column=targ_cols['origem'], value=origem)
        sh.cell(row=j, column=targ_cols['destino'], value=destino)
    else: print(Fore.RED + f'Error reading origem and destino')
    if 'km' in targ_cols:
      print(Fore.YELLOW + 'Reading route kms...')
      for k, row in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
        rota_v = row[rota_i - 1]
        if rota_v is None or not isinstance(rota_v, str):
          print(Fore.RED + 'Error validating value for rota')
          continue
        end_nodes = len(re.findall(r"\w{3}/\w{3}\s", rota_v))
        if end_nodes < 1: continue
        prev_km_cell = sh.cell(row=k, column=targ_cols['km'])
        prev_km = prev_km_cell.value
        if prev_km in ['', None, -1]: 
          sh.cell(row=k, column=targ_cols['km'], value='Não acessível')
          continue
        if isinstance(prev_km, str): prev_km = re.sub(r"\D", '', prev_km) 
        if not isinstance(prev_km, float): prev_km = float(prev_km)
        new_km = prev_km * end_nodes
        sh.cell(row=k, column=targ_cols['km'], value=new_km)
    else: print(Fore.RED + f'Error reading km')
    name, ext = os.path.splitext(f)
    new_f = f'{name}_FILLED{ext}'
    try:
      workbook.save(new_f)
      print(Fore.GREEN + f'Workbook saved successfully as {new_f}')
    except PermissionError:
      print(Fore.RED + f'Permission denied: Unable to save {new_f}. Please ensure the file is not open in another application.')
    except Exception as e:
      print(Fore.RED + f'An error occurred while saving the workbook: {e}')
acc = 0
while acc < 10:
  file_path = input('Enter the path to the Spreadsheet file:\n')
  if os.path.isfile(file_path) and file_path.endswith(('.xls', '.xlsx')):
    print(Fore.GREEN + 'Starting filling procedures...')
    iterate(file_path)
    break
  else:
    print(Fore.YELLOW + 'Invalid file path. Please enter a valid path.')
    acc += 1
    if acc == 10:
      print(Fore.RED + 'Exceeded maximum attempts. Exiting progam.')
      sys.exit()