import json
import os
import sys
import subprocess
import importlib
import platform
from datetime import datetime
from colorama import Fore, init
init(autoreset=True)
perm_error = f"Permission denied by {sys.platform} {platform.platform()} {platform.release()}. Recheck your permission."
wants_xls = False
def prepare(package: str) -> bool:
    """
    Check if 'package' is importable. If not, prompt the user to install.
    Returns True if the package is importable (or installed successfully),
    otherwise False.
    """
    try:
        importlib.import_module(package)
        return True
    except ImportError:
        y = input(
            f'{package} is necessary and not installed in the system. '
            f'Prompt "Y" if you want to install it, else the script execution will be aborted.'
        ).strip().lower()
        if not y:
            return False
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
            return True
        except PermissionError:
            print(perm_error)
            sys.exit(1)
        except Exception as e:
            print('Error log:\n', e)
            print(f"Failed to install {package}. Exiting.")
            sys.exit(2)
if __name__ == '__main__':
    all_prepared = []
    for package in ["pandas", "colorama", "openpyxl"]:
        all_prepared.append(prepare(package))
    if False in all_prepared:
        print("Could not proceed with execution due to missing dependencies or an undefined error. Exiting")
        sys.exit(3)
    if False not in all_prepared: 
        from colorama import Fore, init
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        init(autoreset=True)
        def is_scalar(value) -> bool:
            """Returns True if value is a scalar (atomic) type."""
            return isinstance(value, (str, int, float, bool)) or value is None
        
        def json_prop_to_csv(data, output_prefix: str, dist_path: str) -> None:
            """
            Converts a data object (expected to be a dict or a list of scalars)
            into a CSV file and optionally into an Excel file.
            Only scalar values should be present.
            """
            global wants_xls
            start_time = datetime.now()
            output_name = output_prefix
            try:
                if isinstance(data, dict):
                    df = pd.DataFrame([data])
                elif isinstance(data, list):
                    if all(is_scalar(item) for item in data):
                        df = pd.DataFrame({"value": data})
                    else:
                        df = pd.DataFrame(data)
                else:
                    print(Fore.YELLOW + f"Skipping conversion for {output_name} - not suitable for CSV format")
                    return
                if df.empty:
                    print(Fore.YELLOW + f"Skipping empty data for {output_name}")
                    return
                csv_folder = os.path.join(dist_path, "csv")
                os.makedirs(csv_folder, exist_ok=True)
                output_csv = os.path.join(csv_folder, f"{output_name}.csv")
                df.to_csv(output_csv, index=False, encoding='utf-8', date_format="%d/%m/%y")
                print(Fore.GREEN + f"Successfully converted to CSV: {output_csv} in {(datetime.now() - start_time).total_seconds()} seconds")
            except Exception as e:
                print(Fore.RED + f"Error during CSV conversion for '{output_name}': {str(e)}")
                return
            if not wants_xls:
                wants_xls = input(Fore.CYAN + "Prompt 'Y' if you want to proceed with creating an Excel file:\n").strip().lower() == 'y'
            if not wants_xls:
                return
            try:
                xls_folder = os.path.join(dist_path, "xls")
                os.makedirs(xls_folder, exist_ok=True)
                output_xls = os.path.join(xls_folder, f"{output_prefix}.xlsx")
                sheet_name = os.path.basename(output_xls).split('.')[0][:31]
                to_excel_args = {'engine': 'openpyxl', 'sheet_name': sheet_name, 'index': False}
                os.makedirs(xls_folder, exist_ok=True)
                df.to_excel(output_xls, **to_excel_args)
                wb = load_workbook(output_xls)
                ws = wb.active
                ws.freeze_panes = 'A2'
                thick_black_side = Side(border_style="thick", color="000000")
                thick_greyish_side = Side(border_style="thick", color="191919")
                even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                odd_row_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                dimensions = {}
                for c_i, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=c_i)
                    cell.fill = PatternFill(start_color="FFFFD1", end_color="FFFFD1", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(top=thick_black_side, bottom=thick_black_side)
                    dimensions[c_i] = len(str(col_name))
                for r_i, row in enumerate(df.values, 2):
                    row_fill = even_row_fill if r_i % 2 == 0 else odd_row_fill
                    for c_i, value in enumerate(row, 1):
                        cell = ws.cell(row=r_i, column=c_i)
                        cell.fill = row_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(top=thick_greyish_side, bottom=thick_greyish_side)
                        dimensions[c_i] = max(dimensions.get(c_i, 0), len(str(value)))
                for col, wdt in dimensions.items():
                    ws.column_dimensions[get_column_letter(col)].width = wdt + 4
                wb.save(output_xls)
                print(Fore.GREEN + f"Successfully created Excel file: {output_xls} in {(datetime.now() - start_time).total_seconds()} seconds")
            except Exception as e:
                print(Fore.RED + f"Error during Excel creation for '{output_prefix}': {str(e)}")

        def recurse_level(obj, prefix: str) -> None:
            """
            Recursively walk the JSON object.
            - If obj is a dict, process the scalar key-value pairs (only scalars) and output a CSV.
            - Then, for any non-scalar value, recurse with an updated prefix.
            - If obj is a list (or similar), if all items are scalar, output them; otherwise, process each element.
            """
            print(Fore.BLUE + f"Walking through {prefix}")
            os.makedirs(os.path.dirname(prefix), exist_ok=True)
            if isinstance(obj, dict):
                scalars = {}
                non_scalars = {}
                for k, v in obj.items():
                    if is_scalar(v):
                        scalars[k] = v
                    elif v:
                        non_scalars[k] = v
                if scalars:
                    json_prop_to_csv(scalars, os.path.basename(prefix), os.path.dirname(prefix))
                for key, value in non_scalars.items():
                    new_prefix = os.path.join(prefix, key)
                    recurse_level(value, new_prefix)
            elif isinstance(obj, (list, tuple, set, frozenset)):
                scalars = []
                non_scalars = []
                for item in obj:
                    if is_scalar(item):
                        scalars.append(item)
                    else:
                        non_scalars.append(item)
                if scalars:
                    json_prop_to_csv(scalars, os.path.basename(prefix), os.path.dirname(prefix))
                for idx, item in enumerate(non_scalars):
                    new_prefix = os.path.join(prefix, str(idx+1))
                    recurse_level(item, new_prefix)
            else:
                json_prop_to_csv(obj, os.path.basename(prefix), os.path.dirname(prefix))

        def load_json() -> None:
            """
            Loads a JSON file and processes it recursively to create CSV and Excel files,
            ensuring that only scalar values end up in cells.
            """
            file_path = ''
            dist_path = ''
            acc = 0
            acc2 = 0 
            tries_error = "You have reached the limit of tries. Try again later. Aborting..."
            while not os.path.isfile(file_path):
                try:
                    acc += 1
                    if acc > 10:
                        print(Fore.RED + tries_error)
                        return
                    jsons = []
                    try:
                        for r, _, fs in os.walk(os.getcwd()):
                            for f in fs:
                                if f.endswith('.json'):
                                    jsons.append(os.path.join(os.path.relpath(r), f))
                    except Exception as e:
                        print(Fore.YELLOW + "Could not read paths to jsons")
                    file_path = input(Fore.CYAN + f"Input the relative path to the .json\nAvailable paths are:\n{jsons}\n\nInput the relative path (ignoring initial dots or slashes, and considering your OS' pattern for slashing):\n")
                    if not os.path.isfile(file_path):
                        file_path = ''
                        print(Fore.YELLOW + "The given path is not a file. Please try again.")
                        continue
                    if not file_path.endswith('.json'):
                        file_path = ''
                        print(Fore.YELLOW + "Only .json files are accepted. Please try again.")
                        continue
                    print(Fore.GREEN + f"Starting conversions to .csv at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}...")
                except Exception as e:
                    print(Fore.RED + f"Error: {str(e)}")
                    file_path = ''
                    return
            while not os.path.isdir(dist_path):
                try:
                    acc2 += 1
                    if acc2 > 10:
                        print(Fore.RED + tries_error)
                        return
                    dist_path = input(Fore.CYAN + "Input the destination path to output the CSV/Excel files:\n")
                    if not os.path.isdir(dist_path):
                        d = input(Fore.YELLOW + "The given path is not a directory.\nInput 'd' to use the current working directory or try another relative path:\n").strip().lower() == 'd'
                        if d:
                            dist_path = os.getcwd()
                            break
                        dist_path = ''
                        continue
                    print(Fore.GREEN + "Destination accepted. Proceeding...")
                except Exception as e:
                    print(Fore.RED + f"Error: {str(e)}")
                    dist_path = ''
                    return
            try:
                with open(file_path, mode='r', encoding='utf-8') as f:
                    data = json.load(f)
                print(Fore.BLUE + "Loaded JSON file. Proceeding with conversions...")
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                output_base = os.path.join(dist_path, "tables")
                os.makedirs(output_base, exist_ok=True)
                recurse_level(data, output_base)
            except json.JSONDecodeError as e:
                print(Fore.RED + f"Error decoding JSON file: {str(e)}")
                return
            except Exception as e:
                print(Fore.RED + f"Error during conversion process: {str(e)}")
                return

        load_json()
