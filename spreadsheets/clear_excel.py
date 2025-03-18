import os
import csv
import sys
import subprocess
import time
import openpyxl as op
from datetime import datetime
from collections import defaultdict
from colorama import Fore, init
init(autoreset=True)
def clear_spreadsheets(dirc: str, id_i:int=0) -> None:
    start = datetime.now()
    ss_xl_f = []
    ss_csv_f = []
    print('Starting to populate list of files...')
    # Walk through the directory and collect Excel and CSV files
    for r, _, fs in os.walk(dirc):
        for f in fs:
            if f.endswith('.xlsx') or f.endswith('.xls'):
                ss_xl_f.append(os.path.join(r, f))
            elif f.endswith('.csv'):
                ss_csv_f.append(os.path.join(r, f))
    # Process Excel files
    for x in ss_xl_f:
        try:
            wb = op.load_workbook(x)
            st = wb.active
            d_x = {
                'rows': defaultdict(dict),
                'cols': defaultdict(dict)
            }
            p_ks_x = set()
            for i, r_x in enumerate(st.iter_rows(values_only=True)):
                k_x = r_x[id_i]
                if k_x is None or k_x in p_ks_x:
                  continue
                p_ks_x.add(k_x)
                for j, v_x in enumerate(r_x):
                  d_x['rows'][i][j] = v_x
                  d_x['cols'][j][i] = v_x
                  d_x['rows'][i]['position'] = f'{op.utils.get_column_letter(j+1)}{i+1}'
                  d_x['cols'][j]['position'] = f'{op.utils.get_column_letter(j+1)}{i+1}'
            # Sort the rows alphabetically based on the id_i column
            srtd_x = sorted(d_x['rows'], key=lambda k: (d_x['rows'][k][id_i], k))
            st.delete_rows(1, st.max_row)
            for row_num, sorted_row in enumerate(srtd_x, start=1):
                for col_num, cell_value in d_x['rows'][sorted_row].items():
                  if isinstance(col_num, int):
                    st.cell(row=row_num, column=col_num + 1, value=cell_value)
            new_x_path = x.replace('.xls', '__filtered.xls') if x.endswith('xls') else x.replace('.xlsx', '__filtered.xlsx')
            wb.save(new_x_path)
            print(Fore.BLUE + f'Sucessfully created filtered and sorted file: {new_x_path}')
        except Exception as e:
            print(Fore.RED + f'Error generating Excel file:\n{e}')
    # Process CSV files
    for c in ss_csv_f:
        try:
            d_c = {
                'rows': defaultdict(dict),
                'cols': defaultdict(dict)
            }
            with open(c, newline='', encoding='utf-8') as csv_f:
                csv_rd = csv.reader(csv_f)
                p_ks_c = set()
                for a, r_c in enumerate(csv_rd):
                    k_c = r_c[id_i]
                    if k_c is None or k_c in p_ks_c:
                      continue
                    p_ks_c.add(k_c)
                    for b, v_x in enumerate(r_c):
                      d_c['rows'][a][b] = v_x
                      d_c['cols'][b][a] = v_x
                      d_c['rows'][a]['position'] = f'{chr(65 + b)}{a+1}'
                      d_c['cols'][b]['position'] = f'{chr(65 + b)}{a+1}'
            # Insertion sort for sorting rows alphabetically
            srtd_c = sorted(d_c['rows'], key=lambda k: (d_c['rows'][k][id_i], k))
            new_c_path = c.replace('.csv', '__filtered.csv')
            with open(new_c_path, 'w', newline='', encoding='utf-8') as new_c:
              wt = csv.writer(new_c)
              for _, c_r in srtd_c:
                wt.writerow([c_r[col] for col in sorted(c_r.keys())])
            print(Fore.BLUE + f'Sucessfully filtered and sortes new csv: {new_c_path}')
        except Exception as e:
            print(Fore.RED + f'Error generating CSV file:\n{e}')
    finish = datetime.now()
    print(Fore.GREEN + f'Finished reading spreadsheets in {finish - start}')
def exec_script() -> None:
    try:
        py_v = sys.version
        if not py_v:
            print(Fore.YELLOW + 'Python is not installed on your system.')
            if os.name == 'posix':
                subprocess.run(['sudo', 'apt-get', 'install', 'python3'], check=True)
            elif os.name == 'nt':
                print(Fore.RED + 'Automatic installation of Python is not permitted on Windows Operational System. Exiting...')
                time.sleep(3)
                sys.exit(1)
            else: 
                print(Fore.RED + 'OS not recognized for Python Installion.')
        else:
            print(Fore.BLUE + f'Python is correctly installed. Version: {py_v}. Proceeding to packages check...')
            time.sleep(1)
    except Exception as e:
        print(Fore.RED + f'Error checking or installing Python: {e}')
        sys.exit(1)
    rq_pks = ['openpyxl', 'colorama']
    for pk in rq_pks:
        try:
            __import__(pk)
        except ImportError:
            try:
                print(Fore.YELLOW + f'Package {pk} is missing. Trying installation...')
                subprocess.check_call([sys.executable, '-m', 'pip', 'install', pk])
                print(Fore.GREEN + f'Package {pk} installed sucessfully.')
            except Exception as e:
                print(Fore.RED + f'Failed to install {pk}:\n{e}')
                sys.exit(1)
    key_column = None
    acc = 0
    while acc < 10:
        try:
            key_column = int(input('Enter the index of the identificador column:\n'))
            clear_spreadsheets(os.getcwd(), key_column)
            break
        except ValueError:
            print(Fore.YELLOW + "Invalid input, please enter an integer.")
            acc += 1
        if acc >= 10:
            print(Fore.RED + "Maximum attempts reached, exiting.")
            return
if __name__ == '__main__':
    exec_script()